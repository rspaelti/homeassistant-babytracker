"""Lädt die WHO-Growth-Standards (LMS-Parameter) in die DB.

Quelle: WHO Child Growth Standards, https://www.who.int/childgrowth
Lizenz: CC BY-NC-SA 3.0 IGO
Format: Expanded Tables (Tag-Auflösung) als XLSX
"""

from __future__ import annotations

import argparse
import sys
import urllib.request
from pathlib import Path

from openpyxl import load_workbook
from sqlmodel import Session, select

from babytracker.config import settings
from babytracker.db import engine
from babytracker.models import WhoLms

WHO_BASE = "https://cdn.who.int/media/docs/default-source/child-growth/child-growth-standards/indicators"

# (indicator-name in DB, sex, filename, URL)
FILES: list[tuple[str, str, str, str]] = [
    ("weight", "f", "wfa-girls.xlsx", f"{WHO_BASE}/weight-for-age/expanded-tables/wfa-girls-zscore-expanded-tables.xlsx"),
    ("weight", "m", "wfa-boys.xlsx", f"{WHO_BASE}/weight-for-age/expanded-tables/wfa-boys-zscore-expanded-tables.xlsx"),
    ("length", "f", "lhfa-girls.xlsx", f"{WHO_BASE}/length-height-for-age/expandable-tables/lhfa-girls-zscore-expanded-tables.xlsx"),
    ("length", "m", "lhfa-boys.xlsx", f"{WHO_BASE}/length-height-for-age/expandable-tables/lhfa-boys-zscore-expanded-tables.xlsx"),
    ("head", "f", "hcfa-girls.xlsx", f"{WHO_BASE}/head-circumference-for-age/expanded-tables/hcfa-girls-zscore-expanded-tables.xlsx"),
    ("head", "m", "hcfa-boys.xlsx", f"{WHO_BASE}/head-circumference-for-age/expanded-tables/hcfa-boys-zscore-expanded-tables.xlsx"),
]


def download(url: str, dest: Path) -> None:
    dest.parent.mkdir(parents=True, exist_ok=True)
    if dest.exists() and dest.stat().st_size > 10_000:
        print(f"  bereits vorhanden: {dest.name}")
        return
    print(f"  lade: {url}")
    req = urllib.request.Request(url, headers={"User-Agent": "babytracker/0.1"})
    with urllib.request.urlopen(req, timeout=30) as resp, dest.open("wb") as f:
        f.write(resp.read())


def parse_xlsx(path: Path) -> list[tuple[int, float, float, float]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    rows: list[tuple[int, float, float, float]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c).strip().lower() if c else "" for c in row]
            if headers[:4] != ["day", "l", "m", "s"]:
                raise ValueError(f"Unerwartete Header in {path.name}: {headers[:4]}")
            continue
        day, L, M, S = row[0], row[1], row[2], row[3]
        if day is None or L is None or M is None or S is None:
            continue
        rows.append((int(day), float(L), float(M), float(S)))
    wb.close()
    return rows


def load_all(who_dir: Path, if_empty: bool = False) -> int:
    with Session(engine) as session:
        if if_empty and session.exec(select(WhoLms).limit(1)).first():
            print("WHO-Daten bereits geladen – übersprungen.")
            return 0

        if not if_empty:
            session.query(WhoLms).delete()  # type: ignore[attr-defined]
            session.commit()

        total = 0
        for indicator, sex, filename, url in FILES:
            path = who_dir / filename
            download(url, path)
            rows = parse_xlsx(path)
            for day, L, M, S in rows:
                session.add(
                    WhoLms(indicator=indicator, sex=sex, age_days=day, L=L, M=M, S=S)
                )
            session.flush()
            total += len(rows)
            print(f"  → {indicator}/{sex}: {len(rows)} Zeilen")

        session.commit()
        print(f"Gesamt: {total} LMS-Datensätze.")
        return total


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--if-empty", action="store_true", help="Nur laden wenn Tabelle leer")
    ap.add_argument("--who-dir", type=Path, default=settings.who_dir)
    args = ap.parse_args()

    try:
        load_all(args.who_dir, if_empty=args.if_empty)
    except urllib.error.URLError as e:
        print(f"Download fehlgeschlagen: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
